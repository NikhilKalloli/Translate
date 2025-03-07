import pandas as pd
from deep_translator import GoogleTranslator
import os
from tqdm import tqdm
import shutil
import time
import subprocess
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from copy import copy

def kill_excel_processes():
    """Kill any running Excel processes"""
    try:
        subprocess.run(['taskkill', '/F', '/IM', 'EXCEL.EXE'], capture_output=True)
        time.sleep(2)  # Give some time for the process to be fully killed
    except:
        pass

def copy_cell_format(source_cell, target_cell):
    """Safely copy cell formatting"""
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)

def safe_translate(translator, text, max_retries=3, delay=1):
    """Safely translate text with retries and delay"""
    for attempt in range(max_retries):
        try:
            translated = translator.translate(text)
            time.sleep(delay)  # Add delay between translations
            return translated
        except Exception as e:
            if attempt == max_retries - 1:
                print(f"\nWarning: Could not translate '{text}' after {max_retries} attempts: {e}")
                return text
            time.sleep(delay * (attempt + 1))  # Exponential backoff
    return text

def translate_excel(input_file, output_file=None, source_lang='pt', target_lang='en'):
    """
    Translate Excel file from source language to target language
    """
    # Convert to absolute path
    input_file = os.path.abspath(input_file)
    
    if not output_file:
        base, ext = os.path.splitext(input_file)
        output_file = f"{base}_translated{ext}"
    
    # Ensure Excel is not running
    kill_excel_processes()
    
    print(f"Input file: {input_file}")
    print(f"Output file will be: {output_file}")
    
    # First create a copy of the original file to preserve macros and formatting
    try:
        if os.path.exists(output_file):
            os.remove(output_file)
            time.sleep(1)  # Give some time for the file system
        shutil.copy2(input_file, output_file)
    except FileNotFoundError:
        print(f"Error: Input file not found: {input_file}")
        return
    except Exception as e:
        print(f"Error copying file: {e}")
        return
    
    print(f"Reading Excel file: {input_file}")
    
    # Initialize translator
    translator = GoogleTranslator(source=source_lang, target=target_lang)
    
    try:
        # Load both workbooks
        wb_source = load_workbook(input_file, keep_vba=True, data_only=True)
        wb_target = load_workbook(output_file, keep_vba=True)
        
        total_sheets = len(wb_source.sheetnames)
        current_sheet = 0
        
        # Process each sheet
        for sheet_name in wb_source.sheetnames:
            current_sheet += 1
            print(f"\nProcessing sheet {current_sheet}/{total_sheets}: {sheet_name}")
            ws_source = wb_source[sheet_name]
            ws_target = wb_target[sheet_name]
            
            # Store original column widths and row heights
            column_widths = {col: ws_source.column_dimensions[col].width 
                           for col in ws_source.column_dimensions}
            row_heights = {row: ws_source.row_dimensions[row].height 
                         for row in ws_source.row_dimensions}
            
            # Get all merged cell ranges and store them
            merged_ranges = []
            for merge_range in ws_source.merged_cells.ranges:
                merged_ranges.append(str(merge_range))
            
            # Unmerge all cells in target
            for merge_range in list(ws_target.merged_cells.ranges):
                ws_target.unmerge_cells(str(merge_range))
            
            # Create a dictionary to store translations
            translations = {}
            
            # Count total cells for progress bar
            total_cells = sum(1 for row in ws_source.rows 
                            for cell in row 
                            if not isinstance(cell, MergedCell) and 
                            isinstance(cell.value, str) and cell.value.strip())
            
            # First pass: collect all text that needs translation
            print("Collecting and translating text...")
            with tqdm(total=total_cells, desc="Progress") as pbar:
                for row in ws_source.rows:
                    for cell in row:
                        if isinstance(cell, MergedCell):
                            continue
                        
                        value = cell.value
                        if isinstance(value, str) and value.strip():
                            if value not in translations:
                                translated = safe_translate(translator, value)
                                translations[value] = translated
                                print(f"\nTranslated: '{value}' -> '{translated}'")
                            pbar.update(1)
            
            # Second pass: apply translations and copy formatting
            print("\nApplying translations and formatting...")
            for row in tqdm(ws_source.rows, desc="Progress"):
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    
                    # Get the corresponding cell in target worksheet
                    target_cell = ws_target.cell(row=cell.row, column=cell.column)
                    
                    # Copy value with translation if needed
                    value = cell.value
                    if isinstance(value, str) and value.strip():
                        target_cell.value = translations.get(value, value)
                    else:
                        target_cell.value = value
                    
                    # Copy formatting
                    copy_cell_format(cell, target_cell)
            
            # Restore merged cells
            print("Restoring merged cells...")
            for merge_range in tqdm(merged_ranges, desc="Progress"):
                try:
                    ws_target.merge_cells(merge_range)
                except Exception as e:
                    print(f"\nWarning: Could not merge cells {merge_range}: {e}")
            
            # Restore column widths and row heights
            print("Restoring column widths and row heights...")
            for col, width in column_widths.items():
                if width is not None:
                    ws_target.column_dimensions[col].width = width
            for row, height in row_heights.items():
                if height is not None:
                    ws_target.row_dimensions[row].height = height
            
            # Save progress after each sheet
            print(f"\nSaving progress for sheet {current_sheet}/{total_sheets}...")
            wb_target.save(output_file)
        
        print("\n✓ Translation completed successfully!")
        print(f"✓ Translated file saved as: {output_file}")
        
    except Exception as e:
        print(f"\nError processing Excel file: {e}")
        import traceback
        traceback.print_exc()
        # Clean up the output file if there was an error
        if os.path.exists(output_file):
            try:
                os.remove(output_file)
                print("Cleaned up incomplete output file")
            except:
                pass

if __name__ == "__main__":
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_file = os.path.join(script_dir, "3 - BRDE - Mapa de Comprovação V31.xlsm")
    translate_excel(input_file)
